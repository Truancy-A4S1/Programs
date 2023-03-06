import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def update_workbook(filename):
    workbook = xl.load_workbook('filename')
    sheet = workbook['Sheet1']

    #cell = sheet['coordinates']
    #cell = sheet['a1']
    #print(cell.value)
    cell = sheet.cell(1,1)

    #max number of rows
    print(sheet.max_row)

    for row in range(2, sheet.max_row+1): #2 -> 3 -> 4.. because row 1 is title 
        cell = sheet.cell(row, 3)
        new_price = cell.value * 0.6
        new_price_cell = sheet.cell(row, sheet.max_row)
        new_price_cell.value = new_price

    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    workbook.save(filename)
